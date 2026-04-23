"""
线索看板后端服务 v2
版本: 4.22.018
- 用户认证和数据 API
- 线索分配功能
"""

from flask import Flask, request, jsonify, session, render_template_string
import yaml, json, os, sqlite3, re
from datetime import datetime
from pathlib import Path


def _clean_val(val):
    """清理字符串值，去除换行和回车，避免破坏 JSON/JS 格式"""
    if val is None:
        return ''
    s = str(val)
    s = s.replace('\r', ' ').replace('\n', ' ')
    return s.strip()

app = Flask(__name__)
app.secret_key = 'xiansuo-kanban-secret-key-2024'

BASE_DIR = Path(__file__).parent
USERS_FILE = BASE_DIR / 'users.yaml'
DATA_FILE = BASE_DIR / 'dashboard_data.json'
DB_FILE = BASE_DIR / 'leads.db'

# ─────────────────────────────────────────────
# 数据库初始化
# ─────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS new_leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            phone TEXT NOT NULL,
            platform TEXT NOT NULL,
            agent TEXT NOT NULL,
            entry_date TEXT DEFAULT '',
            name TEXT DEFAULT '',
            city TEXT DEFAULT '',
            validity TEXT DEFAULT '',
            region TEXT DEFAULT '',
            can_wechat TEXT DEFAULT '',
            remark TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            is_read INTEGER DEFAULT 0,
            xhs_account TEXT DEFAULT '',
            lead_type TEXT DEFAULT '',
            UNIQUE(phone)
        )
    ''')
    # 添加新字段（如果不存在）
    try:
        c.execute('ALTER TABLE new_leads ADD COLUMN xhs_account TEXT DEFAULT ""')
    except:
        pass
    try:
        c.execute('ALTER TABLE new_leads ADD COLUMN lead_type TEXT DEFAULT ""')
    except:
        pass
    conn.commit()
    conn.close()
    # 线索成本表
    c.execute('''
        CREATE TABLE IF NOT EXISTS cost_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cost_date TEXT NOT NULL,
            platform TEXT NOT NULL,
            amount REAL NOT NULL,
            unit_cost REAL DEFAULT 0,
            created_at TEXT NOT NULL,
            UNIQUE(cost_date, platform)
        )
    ''')
    # 兼容旧表：若缺少 unit_cost 列则添加
    try:
        c.execute('SELECT unit_cost FROM cost_data LIMIT 1')
    except sqlite3.OperationalError:
        c.execute('ALTER TABLE cost_data ADD COLUMN unit_cost REAL DEFAULT 0')
    conn.commit()
    conn.close()

init_db()

# 加载用户配置
def load_users():
    with open(USERS_FILE, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

# 加载原始线索数据（自动去重，以手机号为主键）
def load_data():
    if DATA_FILE.exists():
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        # 去重：保留每个手机号的第一条记录
        seen = {}
        result = []
        for r in data:
            phone = str(r.get('手机号', '') or r.get('手机', '')).strip()
            if phone and phone not in seen:
                seen[phone] = True
                result.append(r)
        if len(result) < len(data):
            print(f"[去重] 原始数据 {len(data)} 条，去重后 {len(result)} 条")
        return result
    return []

# 加载新录入线索
def load_new_leads():
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute('SELECT * FROM new_leads ORDER BY created_at DESC')
    rows = c.fetchall()
    conn.close()
    
    leads = []
    for row in rows:
        leads.append({
            'id': row[0],
            '手机号': _clean_val(row[1]),
            '平台': _clean_val(row[2]),
            '所属招商': _clean_val(row[3]),
            '录入日期': _clean_val(row[12]),
            '姓名': _clean_val(row[4]),
            '省份': _clean_val(row[5]),
            '线索有效性': _clean_val(row[6]),
            '所属大区': _clean_val(row[7]),
            '是否能加上微信': _clean_val(row[8]),
            '备注': _clean_val(row[9]),
            '入库时间': _clean_val(row[12]) or (_clean_val(row[10])[:10] if len(row) > 10 else ''),
            '是否已读': row[11] if len(row) > 11 else 0,
            '二次联系时间': _clean_val(row[13]),
            '二次联系备注': _clean_val(row[14]),
            '最近一次电联时间': _clean_val(row[15]),
            '到访时间': _clean_val(row[16]),
            '签约时间': _clean_val(row[17]),
            '小红书账号': _clean_val(row[18]) if len(row) > 18 else '',
            '线索类型': _clean_val(row[19]) if len(row) > 19 else '',
            '来源文件': '手动录入'
        })
    return leads

# 加载成本数据
def load_cost_data():
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute('SELECT cost_date, platform, amount, unit_cost FROM cost_data ORDER BY cost_date ASC')
    rows = c.fetchall()
    conn.close()
    return [{'date': r[0], 'platform': r[1], 'amount': r[2], 'unit_cost': r[3] if r[3] else 0} for r in rows]

# ─────────────────────────────────────────────
# 登录接口
# ─────────────────────────────────────────────
@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    
    if not username or not password:
        return jsonify({'success': False, 'message': '请输入用户名和密码'})
    
    users = load_users()
    
    if username == users['admin']['username'] and password == users['admin']['password']:
        session['user'] = {
            'username': username,
            'name': users['admin']['name'],
            'role': 'admin'
        }
        return jsonify({'success': True, 'name': users['admin']['name'], 'role': 'admin'})
    
    for agent in users.get('agents', []):
        if agent['username'] == username and agent['password'] == password:
            session['user'] = {
                'username': username,
                'name': agent['name'],
                'role': 'agent',
                'regions': agent.get('regions', [])
            }
            return jsonify({'success': True, 'name': agent['name'], 'role': 'agent'})
    
    # 游客账号登录
    guest = users.get('guest')
    if guest and username == guest['username'] and password == guest['password']:
        session['user'] = {
            'username': username,
            'name': guest['name'],
            'role': 'guest'
        }
        return jsonify({'success': True, 'name': guest['name'], 'role': 'guest'})
    
    return jsonify({'success': False, 'message': '用户名或密码错误'})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/current_user', methods=['GET'])
def current_user():
    user = session.get('user')
    if user:
        return jsonify({'logged_in': True, 'user': user})
    return jsonify({'logged_in': False})

# ─────────────────────────────────────────────
# 线索数据 API（合并原始数据 + 新录入线索）
# ─────────────────────────────────────────────
@app.route('/api/leads', methods=['GET'])
def get_leads():
    user = session.get('user')
    if not user:
        return jsonify({'error': '请先登录'}), 401
    
    records = load_data()
    new_leads = load_new_leads()
    
    # 管理员和游客看全部
    if user['role'] == 'admin' or user['role'] == 'guest':
        all_records = records + new_leads
        return jsonify({'records': all_records, 'total': len(all_records), 'new_leads_count': len(new_leads)})
    
    # 招商员只看自己分配的
    agent_name = user['name']
    filtered = [r for r in records if r.get('所属招商', '') == agent_name 
                or r.get('跟进员工', '') == agent_name]
    
    # 加上新录入的线索
    agent_new_leads = [r for r in new_leads if r.get('所属招商', '') == agent_name]
    all_filtered = filtered + agent_new_leads
    
    # 获取未读新线索数
    unread = len([r for r in agent_new_leads if r.get('是否已读', 1) == 0])
    
    return jsonify({
        'records': all_filtered, 
        'total': len(all_filtered), 
        'role': 'agent',
        'new_leads_count': len(agent_new_leads),
        'unread_count': unread
    })

# ─────────────────────────────────────────────
# 新线索录入 API（管理员）
# ─────────────────────────────────────────────
@app.route('/api/leads/add', methods=['POST'])
def add_lead():
    user = session.get('user')
    if not user or user['role'] != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以录入线索'}), 401
    
    data = request.json
    phone = data.get('phone', '').strip()
    platform = data.get('platform', '').strip()
    agent = data.get('agent', '').strip()
    entry_date = data.get('entry_date', '').strip() or datetime.now().strftime('%Y-%m-%d')
    
    if not phone or not platform or not agent:
        return jsonify({'success': False, 'message': '请填写完整信息'})
    
    # 检查是否已存在
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute('SELECT id FROM new_leads WHERE phone = ?', (phone,))
    if c.fetchone():
        conn.close()
        return jsonify({'success': False, 'message': '该手机号已录入'})
    
    # 插入新线索
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    c.execute('''
        INSERT INTO new_leads (phone, platform, agent, entry_date, created_at)
        VALUES (?, ?, ?, ?, ?)
    ''', (phone, platform, agent, entry_date, now))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': '线索录入成功'})

# ─────────────────────────────────────────────
# 线索成本录入 API（管理员）
# ─────────────────────────────────────────────
@app.route('/api/cost/add', methods=['POST'])
def add_cost():
    user = session.get('user')
    if not user or user['role'] != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以录入成本'}), 401
    
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'success': False, 'message': '请求数据为空'})
    except Exception as e:
        return jsonify({'success': False, 'message': '请求解析失败: ' + str(e)})
    
    cost_date = str(data.get('cost_date', '')).strip()
    platform = str(data.get('platform', '')).strip()
    amount = data.get('amount', 0)
    unit_cost = data.get('unit_cost', 0)
    
    if not cost_date or not platform:
        return jsonify({'success': False, 'message': '请填写日期和平台'})
    
    try:
        amount = float(amount)
    except:
        return jsonify({'success': False, 'message': '金额格式错误'})
    
    try:
        unit_cost = float(unit_cost) if unit_cost else 0
    except:
        unit_cost = 0
    
    try:
        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 兼容低版本 SQLite：手动判断更新或插入
        c.execute('SELECT id FROM cost_data WHERE cost_date = ? AND platform = ?', (cost_date, platform))
        existing = c.fetchone()
        if existing:
            c.execute('''
                UPDATE cost_data SET amount = ?, unit_cost = ?, created_at = ? WHERE id = ?
            ''', (amount, unit_cost, now, existing[0]))
        else:
            c.execute('''
                INSERT INTO cost_data (cost_date, platform, amount, unit_cost, created_at)
                VALUES (?, ?, ?, ?, ?)
            ''', (cost_date, platform, amount, unit_cost, now))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': f'{platform} {cost_date} 成本录入成功'})
    except Exception as e:
        return jsonify({'success': False, 'message': '录入失败: ' + str(e)})

# ─────────────────────────────────────────────
# 获取成本数据 API
# ─────────────────────────────────────────────
@app.route('/api/cost', methods=['GET'])
def get_cost():
    user = session.get('user')
    if not user or user['role'] != 'admin':
        return jsonify({'error': '无权限'}), 401
    
    cost_data = load_cost_data()
    return jsonify({'cost_data': cost_data})

# ─────────────────────────────────────────────
# 删除成本记录 API（管理员）
# ─────────────────────────────────────────────
@app.route('/api/cost/delete', methods=['POST'])
def delete_cost():
    user = session.get('user')
    if not user or user['role'] != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以删除成本'}), 401
    
    data = request.json
    cost_date = data.get('cost_date', '').strip()
    platform = data.get('platform', '').strip()
    
    if not cost_date or not platform:
        return jsonify({'success': False, 'message': '请提供日期和平台'})
    
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute('DELETE FROM cost_data WHERE cost_date = ? AND platform = ?', (cost_date, platform))
    deleted = c.rowcount
    conn.commit()
    conn.close()
    
    if deleted > 0:
        return jsonify({'success': True, 'message': f'已删除 {platform} {cost_date} 的成本记录'})
    else:
        return jsonify({'success': False, 'message': '未找到该记录'})

# ─────────────────────────────────────────────
# 招商员更新线索详情
# ─────────────────────────────────────────────
@app.route('/api/leads/update', methods=['POST'])
def update_lead():
    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '请先登录'}), 401
    
    data = request.json
    phone = data.get('phone', '').strip()
    
    # 检查是否是该招商员的线索
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute('SELECT id, agent FROM new_leads WHERE phone = ?', (phone,))
    row = c.fetchone()
    
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': '线索不存在'})
    
    lead_id, lead_agent = row
    
    # 管理员可以更新所有，招商员只能更新自己的
    if user['role'] != 'admin' and lead_agent != user['name']:
        conn.close()
        return jsonify({'success': False, 'message': '无权修改此线索'})

    # 基础字段（所有用户可修改）
    name = data.get('name', '')
    city = data.get('city', '')
    validity = data.get('validity', '')
    region = data.get('region', '')
    can_wechat = data.get('can_wechat', '')
    remark = data.get('remark', '')
    entry_date = data.get('entry_date', '')
    contact_time = data.get('二次联系时间', '')
    contact_remark = data.get('二次联系备注', '')
    call_time = data.get('最近一次电联时间', '')
    visit_time = data.get('到访时间', '')
    sign_time = data.get('签约时间', '')
    platform = data.get('platform', '')
    xhs_account = data.get('xhs_account', '')
    lead_type = data.get('lead_type', '')

    # 管理员可修改平台和入库日期，普通招商员保留原值
    if user['role'] == 'admin':
        update_sql = '''
            UPDATE new_leads SET
                name = ?, city = ?, validity = ?, region = ?, can_wechat = ?, remark = ?,
                platform = ?, entry_date = ?, 二次联系时间 = ?, 二次联系备注 = ?, 最近一次电联时间 = ?, 到访时间 = ?, 签约时间 = ?,
                xhs_account = ?, lead_type = ?
            WHERE id = ?
        '''
        update_params = (name, city, validity, region, can_wechat, remark, platform, entry_date, contact_time, contact_remark, call_time, visit_time, sign_time, xhs_account, lead_type, lead_id)
    else:
        # 保留原平台和入库日期
        c.execute('SELECT platform, entry_date FROM new_leads WHERE id = ?', (lead_id,))
        orig = c.fetchone()
        orig_platform = orig[0] if orig else ''
        orig_entry_date = orig[1] if orig else ''
        update_sql = '''
            UPDATE new_leads SET
                name = ?, city = ?, validity = ?, region = ?, can_wechat = ?, remark = ?,
                二次联系时间 = ?, 二次联系备注 = ?, 最近一次电联时间 = ?, 到访时间 = ?, 签约时间 = ?,
                xhs_account = ?, lead_type = ?
            WHERE id = ?
        '''
        update_params = (name, city, validity, region, can_wechat, remark, contact_time, contact_remark, call_time, visit_time, sign_time, xhs_account, lead_type, lead_id)

    c.execute(update_sql, update_params)
    conn.commit()
    conn.close()

    return jsonify({'success': True, 'message': '更新成功'})

# ─────────────────────────────────────────────
# 删除线索
# ─────────────────────────────────────────────
@app.route('/api/leads/delete', methods=['POST'])
def delete_lead():
    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '请先登录'}), 401

    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'success': False, 'message': '请求数据为空'})
        lead_id = data.get('id')
        phone = str(data.get('phone', '')).strip()
        if not lead_id and not phone:
            return jsonify({'success': False, 'message': '请提供id或手机号'})
    except Exception as e:
        return jsonify({'success': False, 'message': '请求解析失败: ' + str(e)})

    try:
        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()
        if lead_id:
            c.execute('SELECT id, agent FROM new_leads WHERE id = ?', (lead_id,))
        else:
            c.execute('SELECT id, agent FROM new_leads WHERE phone = ?', (phone,))
        row = c.fetchone()

        if not row:
            conn.close()
            return jsonify({'success': False, 'message': '线索不存在或无法删除（仅支持删除手动录入的线索）'})

        lead_id, lead_agent = row

        # 管理员可以删除所有，招商员只能删除自己的
        if user['role'] != 'admin' and lead_agent != user['name']:
            conn.close()
            return jsonify({'success': False, 'message': '无权删除此线索'})

        c.execute('DELETE FROM new_leads WHERE id = ?', (lead_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '删除成功'})
    except Exception as e:
        return jsonify({'success': False, 'message': '删除出错: ' + str(e)})

# ─────────────────────────────────────────────
# 批量删除线索
# ─────────────────────────────────────────────
@app.route('/api/leads/batch-delete', methods=['POST'])
def batch_delete_leads():
    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '请先登录'}), 401

    try:
        data = request.get_json(force=True)
        phones = data.get('phones', [])
        if not phones or not isinstance(phones, list):
            return jsonify({'success': False, 'message': '请提供要删除的手机号列表'})
    except Exception as e:
        return jsonify({'success': False, 'message': '请求解析失败: ' + str(e)})

    try:
        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()

        deleted = 0
        skipped = 0
        for phone in phones:
            phone = str(phone).strip()
            if not phone:
                skipped += 1
                continue

            c.execute('SELECT id, agent FROM new_leads WHERE phone = ?', (phone,))
            row = c.fetchone()
            if not row:
                skipped += 1
                continue

            lead_id, lead_agent = row
            # 权限检查
            if user['role'] != 'admin' and lead_agent != user['name']:
                skipped += 1
                continue

            c.execute('DELETE FROM new_leads WHERE id = ?', (lead_id,))
            deleted += 1

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': f'批量删除完成：成功 {deleted} 条，跳过 {skipped} 条'})
    except Exception as e:
        return jsonify({'success': False, 'message': '删除出错: ' + str(e)})

# ─────────────────────────────────────────────
# 标记新线索为已读
# ─────────────────────────────────────────────
@app.route('/api/leads/mark_read', methods=['POST'])
def mark_lead_read():
    user = session.get('user')
    if not user:
        return jsonify({'success': False, 'message': '请先登录'}), 401
    
    data = request.json
    lead_id = data.get('id')
    
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute('UPDATE new_leads SET is_read = 1 WHERE id = ?', (lead_id,))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

# ─────────────────────────────────────────────
# 获取未读提示
# ─────────────────────────────────────────────
@app.route('/api/notifications', methods=['GET'])
def get_notifications():
    user = session.get('user')
    if not user:
        return jsonify({'error': '请先登录'}), 401
    
    if user['role'] == 'admin':
        new_leads = load_new_leads()
        unread = [l for l in new_leads if l.get('是否已读', 1) == 0]
        return jsonify({'unread_count': len(unread), 'notifications': unread})
    
    agent_name = user['name']
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute('SELECT * FROM new_leads WHERE agent = ? AND is_read = 0 ORDER BY created_at DESC', (agent_name,))
    rows = c.fetchall()
    conn.close()
    
    notifications = []
    for row in rows:
        notifications.append({
            'id': row[0],
            '手机号': row[1],
            '平台': row[2],
            '入库时间': row[10][:10] if row[10] else ''
        })
    
    return jsonify({'unread_count': len(notifications), 'notifications': notifications})

# ─────────────────────────────────────────────
# 导入Excel线索（管理员）— v3 完整字段版
# ─────────────────────────────────────────────
@app.route('/api/leads/import', methods=['POST'])
def import_leads():
    user = session.get('user')
    if not user or user['role'] != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以导入线索'}), 401

    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '请选择文件'})

    try:
        import pandas as pd
        from io import BytesIO

        file_bytes = file.read()
        filename = file.filename
        filename_lower = filename.lower()

        # ── 1) 读取Excel ──
        bio = BytesIO(file_bytes)
        df = None
        try:
            if filename_lower.endswith('.xls'):
                df = pd.read_excel(bio, engine='xlrd')
            else:
                df = pd.read_excel(bio, engine='openpyxl')
        except Exception as read_err:
            err_msg = str(read_err)
            if 'not supported' in err_msg.lower():
                err_msg += '（服务器缺少openpyxl库，请联系管理员安装: pip install openpyxl）'
            return jsonify({'success': False, 'message': f'读取Excel失败: {err_msg}'})

        # 兼容 pandas 1.1.5：直接读取异常时回退到 openpyxl 手动构造
        if not filename_lower.endswith('.xls') and (df is None or len(df) == 0 or len(df.columns) <= 1):
            try:
                bio.seek(0)
                from openpyxl import load_workbook
                wb = load_workbook(bio, data_only=True)
                ws = wb.active
                data = []
                headers = None
                for row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(c).strip() if c is not None else '' for c in row]
                    else:
                        data.append(row)
                df = pd.DataFrame(data, columns=headers)
                if df is not None and len(df) > 0:
                    df = df.dropna(how='all')
            except Exception:
                pass

        if df is None or len(df) == 0:
            return jsonify({'success': False, 'message': 'Excel 文件为空（0行数据）'})

        cols = [str(c).strip() for c in df.columns]
        print(f"[导入] 文件={filename}, 行数={len(df)}, 列名={cols}")

        # ── 2) 智能识别列 ──
        def find_col(keywords):
            for kw in keywords:
                kw_lower = kw.lower()
                for col in cols:
                    if kw_lower in col.lower():
                        return col
            return None

        phone_col    = find_col(['手机号', '手机号码', '电话', '联系电话', '手机', 'phone', '客户电话'])
        platform_col = find_col(['平台', '来源', '渠道', '来源平台', '线索来源'])
        agent_col    = find_col(['所属招商', '跟进员工', '负责人', '招商员', '员工', '分配'])
        date_col     = find_col(['入库日期', '录入日期', '日期', '入库时间', '录入时间'])
        name_col     = find_col(['姓名', '名字', '客户姓名', '联系人'])
        city_col     = find_col(['城市', '省份', '地区', '所在城市', '省'])
        region_col   = find_col(['所属大区', '大区', '区域'])
        validity_col = find_col(['线索有效性', '有效性', '客户类型', '等级'])
        wechat_col   = find_col(['是否能加上微信', '能否加微', '加微信', '微信'])
        remark_col   = find_col(['备注', '说明', '备注信息', '客户情况备注'])
        follow_time_col  = find_col(['二次联系时间', '跟进时间', '下次联系时间'])
        follow_note_col  = find_col(['二次联系备注', '跟进备注', '联系备注'])
        call_time_col    = find_col(['最近一次电联时间', '电联时间', '最近联系时间'])
        visit_time_col   = find_col(['到访时间', '到店时间', '来访时间'])
        sign_time_col    = find_col(['签约时间', '成交时间', '签约日期'])

        if not phone_col:
            return jsonify({'success': False,
                'message': f'无法识别手机号列。当前列名: {cols}'})

        print(f"[导入] 列映射: 手机={phone_col}, 平台={platform_col}, 招商={agent_col}, 日期={date_col}")

        # ── 3) 文件类型判断 ──
        req_type = request.form.get('type', '')
        is_zhaoshang = '招商' in filename or req_type == 'zhaoshang'
        is_douyin_channel = req_type == 'douyin'
        is_xhs_channel = req_type == 'xiaohongshu'
        is_douyin_kezi = '客资' in filename or ('抖音' in filename and not is_zhaoshang)

        # ── 4) 解析所有行 ──
        def get_val(row, col):
            if not col:
                return ''
            v = row.get(col, '')
            if pd.isna(v):
                return ''
            s = str(v).strip()
            return s if s.lower() != 'nan' else ''

        def parse_date(row, col):
            if not col:
                return ''
            v = row.get(col)
            if pd.isna(v):
                return ''
            try:
                return pd.to_datetime(v).strftime('%Y-%m-%d')
            except:
                pass
            # 兼容中文日期格式：2026年4月14日
            s = str(v).strip()
            m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', s)
            if m:
                return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
            return ''

        parsed = []
        bad_rows = []
        for idx, row in df.iterrows():
            raw = row.get(phone_col, '')
            if pd.isna(raw):
                bad_rows.append({'row': int(idx)+2, 'raw': '空值', 'reason': '联系方式为空'})
                continue
            s = str(raw).strip()
            if not s or s.lower() == 'nan':
                bad_rows.append({'row': int(idx)+2, 'raw': '空值', 'reason': '联系方式为空'})
                continue

            digits = ''.join(filter(str.isdigit, s))
            if len(digits) == 11:
                phone = digits
            elif len(digits) >= 7:
                phone = digits
            else:
                phone = s  # 微信号保留

            # 入库日期：以Excel表格为准
            entry_date = parse_date(row, date_col)
            if not entry_date:
                for fc in ['日期', '时间', '创建日期', '添加日期']:
                    col = find_col([fc])
                    if col:
                        entry_date = parse_date(row, col)
                        if entry_date:
                            break
            if not entry_date:
                entry_date = datetime.now().strftime('%Y-%m-%d')

            # 招商员分配
            if is_douyin_kezi:
                agent = get_val(row, agent_col) or '郑建军'
            else:
                agent = get_val(row, find_col(['所属招商'])) or get_val(row, find_col(['跟进员工'])) or get_val(row, agent_col) or '郑建军'

            parsed.append({
                'phone': phone,
                'platform': '小红书' if is_xhs_channel else ('抖音' if is_douyin_channel else (get_val(row, platform_col) or '抖音')),
                'agent': agent,
                'entry_date': entry_date,
                'name': get_val(row, name_col),
                'city': get_val(row, city_col),
                'region': get_val(row, region_col),
                'validity': get_val(row, validity_col),
                'can_wechat': get_val(row, wechat_col),
                'remark': get_val(row, remark_col),
                'follow_time': parse_date(row, follow_time_col),
                'follow_note': get_val(row, follow_note_col),
                'call_time': parse_date(row, call_time_col),
                'visit_time': parse_date(row, visit_time_col),
                'sign_time': parse_date(row, sign_time_col),
            })

        if not parsed:
            return jsonify({'success': False, 'message': '未能解析出任何有效数据'})

        print(f"[导入] 解析成功: {len(parsed)} 条, 空值跳过: {len(bad_rows)} 条")

        # ── 5) 数据库写入 ──
        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()

        # 5.1) 清理已有重复
        c.execute('SELECT phone, COUNT(*) as cnt FROM new_leads GROUP BY phone HAVING cnt > 1')
        for phone, cnt in c.fetchall():
            c.execute('DELETE FROM new_leads WHERE phone = ? AND id NOT IN (SELECT MIN(id) FROM new_leads WHERE phone = ?)', (phone, phone))
        conn.commit()

        # 5.2) 加载已有线索
        c.execute('SELECT phone, platform, entry_date FROM new_leads')
        existing = {row[0]: {'platform': row[1], 'entry_date': row[2]} for row in c.fetchall()}

        added, updated, dup_skip, skipped = 0, 0, 0, 0
        seen_in_file = set()

        for item in parsed:
            phone = item['phone']
            if phone in seen_in_file:
                dup_skip += 1
                continue
            seen_in_file.add(phone)

            platform = item['platform']
            agent = item['agent']
            entry_date = item['entry_date']

            if phone in existing:
                old = existing[phone]
                # 抖音/小红书渠道导入时，已存在的线索直接跳过（由管理表维护）
                if is_douyin_channel or is_xhs_channel:
                    skipped += 1
                    continue

                # 招商线索管理表导入时，抖音/小红书入库日期不变（兼容"抖音广告"等变体）
                if is_zhaoshang and ('抖音' in old['platform'] or '小红书' in old['platform']):
                    entry_date = old['entry_date']

                # UPDATE 全部字段（region 加入更新）
                c.execute('''UPDATE new_leads SET
                    platform = ?, agent = ?, entry_date = ?, name = ?,
                    city = ?, region = ?, validity = ?, can_wechat = ?, remark = ?,
                    二次联系时间 = ?, 二次联系备注 = ?, 最近一次电联时间 = ?, 到访时间 = ?, 签约时间 = ?
                    WHERE phone = ?''', (
                    platform, agent, entry_date, item['name'],
                    item['city'], item['region'], item['validity'], item['can_wechat'], item['remark'],
                    item['follow_time'], item['follow_note'], item['call_time'],
                    item['visit_time'], item['sign_time'], phone))
                updated += 1
            else:
                # 招商线索管理表导入时，抖音/小红书的新线索不创建（由渠道表格单独导入）
                if is_zhaoshang and ('抖音' in platform or '小红书' in platform):
                    skipped += 1
                    continue

                # INSERT 新线索（region 加入插入）
                c.execute('''INSERT INTO new_leads
                    (phone, platform, agent, entry_date, name, city, region, validity,
                     can_wechat, remark, created_at,
                     二次联系时间, 二次联系备注, 最近一次电联时间, 到访时间, 签约时间)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
                    phone, platform, agent, entry_date, item['name'], item['city'], item['region'],
                    item['validity'], item['can_wechat'], item['remark'],
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    item['follow_time'], item['follow_note'], item['call_time'],
                    item['visit_time'], item['sign_time']))
                added += 1
                existing[phone] = {'platform': platform, 'entry_date': entry_date}

        conn.commit()
        conn.close()

        msg = f'导入完成！新增 {added} 条，更新 {updated} 条'
        if skipped:
            if is_douyin_channel or is_xhs_channel:
                msg += f'，已存在线索跳过 {skipped} 条'
            else:
                msg += f'，抖音/小红书新线索跳过 {skipped} 条（请通过渠道表格导入）'
        if dup_skip:
            msg += f'，Excel内重复跳过 {dup_skip} 条'
        if bad_rows:
            msg += f'，空值跳过 {len(bad_rows)} 条'

        return jsonify({
            'success': True, 'message': msg,
            'added': added, 'updated': updated,
            'dup_skip': dup_skip, 'bad': len(bad_rows),
            'bad_rows': bad_rows[:5]
        })

    except Exception as e:
        import traceback
        print(f"[导入错误] {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})

# ─────────────────────────────────────────────
# 抖音客资批量导入（管理员）
# ─────────────────────────────────────────────
@app.route('/api/leads/import-douyin', methods=['POST'])
def import_douyin_kezi():
    user = session.get('user')
    if not user or user['role'] != 'admin':
        return jsonify({'success': False, 'message': '只有管理员可以导入'}), 401

    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '请选择文件'})

    try:
        import pandas as pd
        from io import BytesIO
        from openpyxl import load_workbook

        file_bytes = file.read()
        bio = BytesIO(file_bytes)
        filename_lower = file.filename.lower()
        is_xls = filename_lower.endswith('.xls')

        # 获取 sheet 名称列表
        if is_xls:
            xls = pd.ExcelFile(bio, engine='xlrd')
            sheet_names = xls.sheet_names
        else:
            bio.seek(0)
            wb = load_workbook(bio, data_only=True)
            sheet_names = wb.sheetnames

        def load_sheet(bio, filename, sheet_name=None):
            """统一读取Excel返回DataFrame。xlsx走openpyxl，xls走pandas+xlrd"""
            if filename.lower().endswith('.xls'):
                bio.seek(0)
                if sheet_name:
                    df = pd.read_excel(bio, sheet_name=sheet_name, engine='xlrd')
                else:
                    df = pd.read_excel(bio, engine='xlrd')
            else:
                bio.seek(0)
                wb = load_workbook(bio, data_only=True)
                ws = wb[sheet_name] if sheet_name else wb.active
                data = []
                headers = None
                for row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(c).strip() if c is not None else '' for c in row]
                    else:
                        data.append(row)
                df = pd.DataFrame(data, columns=headers)
            if df is not None and len(df) > 0:
                df = df.dropna(how='all')
            return df

        # 直接读取第一个有数据的sheet
        df = None
        debug_info = []
        try:
            df = load_sheet(bio, file.filename)
            debug_info.append(f"直接读取成功，行数={len(df)}, 列数={len(df.columns)}")
        except Exception as e:
            debug_info.append(f"直接读取失败: {e}")

        # 如果直接读取为空，逐个sheet尝试
        if df is None or len(df) == 0:
            for sheet_name in sheet_names:
                try:
                    df = load_sheet(bio, file.filename, sheet_name)
                    debug_info.append(f"Sheet {sheet_name} 读取成功，行数={len(df)}")
                    if len(df) > 0:
                        break
                except Exception as e2:
                    debug_info.append(f"Sheet {sheet_name} 读取失败: {e2}")
                    continue

        if df is None or len(df) == 0:
            return jsonify({'success': False, 'message': f'Excel 文件为空。调试: {" | ".join(debug_info)}'})

        cols = [str(c).strip() for c in df.columns]

        # 查找对应列
        def find_col(keywords):
            for kw in keywords:
                kw_lower = kw.lower()
                for col in cols:
                    if kw_lower in col.lower():
                        return col
            return None

        date_col   = find_col(['线索创建时间', '创建时间', '日期', '时间'])
        name_col   = find_col(['姓名', '名字', '客户姓名'])
        phone_col  = find_col(['手机号', '手机号码', '电话', '联系电话', '手机'])
        agent_col  = find_col(['跟进员工', '所属招商', '负责人', '招商员', '员工'])
        city_col   = find_col(['所在城市', '城市', '省份', '地区'])

        if not phone_col:
            return jsonify({'success': False, 'message': f'无法识别手机号列。当前列名: {cols}'})

        # 加载已有手机号（用于去重）
        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()
        c.execute('SELECT phone FROM new_leads')
        existing_phones = {row[0] for row in c.fetchall()}
        conn.close()

        def get_val(row, col):
            if not col:
                return ''
            v = row.get(col, '')
            if pd.isna(v):
                return ''
            s = str(v).strip()
            return s if s.lower() != 'nan' else ''

        def parse_date(row, col):
            if not col:
                return ''
            v = row.get(col)
            if pd.isna(v):
                return ''
            try:
                return pd.to_datetime(v).strftime('%Y-%m-%d')
            except:
                return ''

        added, skipped, bad = 0, 0, 0
        seen = set()

        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()

        for idx, row in df.iterrows():
            raw_phone = get_val(row, phone_col)
            if not raw_phone:
                bad += 1
                continue

            # 提取数字
            digits = ''.join(filter(str.isdigit, raw_phone))
            if len(digits) >= 7:
                phone = digits
            else:
                phone = raw_phone

            # 去重：数据库已有 或 Excel内重复
            if phone in existing_phones or phone in seen:
                skipped += 1
                continue
            seen.add(phone)

            entry_date = parse_date(row, date_col) or datetime.now().strftime('%Y-%m-%d')
            name = get_val(row, name_col)
            agent = get_val(row, agent_col) or '郑建军'
            city = get_val(row, city_col)

            c.execute('''INSERT INTO new_leads
                (phone, platform, agent, entry_date, name, city, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)''', (
                phone, '抖音', agent, entry_date, name, city,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ))
            added += 1

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': f'导入完成！新增 {added} 条，重复跳过 {skipped} 条，无法识别 {bad} 条',
            'added': added,
            'skipped': skipped,
            'bad': bad
        })

    except Exception as e:
        import traceback
        print(f"[抖音导入错误] {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})

# ─────────────────────────────────────────────
# 管理员页面（包含线索录入表单）
# ─────────────────────────────────────────────
@app.route('/admin', methods=['GET'])
def admin_page():
    user = session.get('user')
    if not user or user['role'] != 'admin':
        return '<h1>只有管理员可以访问</h1><a href="/">返回</a>'
    
    # 获取招商员列表
    users = load_users()
    agents = [a['name'] for a in users.get('agents', [])]
    
    html = '''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>线索录入 - 管理后台</title>
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            body { font-family: -apple-system, BlinkMacSystemFont, sans-serif; background: #f5f7fa; padding: 20px; }
            .header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 12px; margin-bottom: 20px; display: flex; justify-content: space-between; align-items: center; }
            .header h1 { font-size: 24px; }
            .nav { display: flex; gap: 15px; }
            .nav a { color: white; text-decoration: none; padding: 8px 16px; background: rgba(255,255,255,0.2); border-radius: 8px; }
            .nav a:hover { background: rgba(255,255,255,0.3); }
            .container { background: white; border-radius: 12px; padding: 30px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
            h2 { margin-bottom: 20px; color: #333; }
            .form { display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; max-width: 800px; }
            .form-group { display: flex; flex-direction: column; }
            .form-group label { font-size: 14px; color: #666; margin-bottom: 5px; }
            .form-group input, .form-group select { padding: 10px; border: 2px solid #e0e0e0; border-radius: 8px; font-size: 14px; }
            .form-group input:focus, .form-group select:focus { outline: none; border-color: #667eea; }
            .btn { grid-column: 1 / -1; padding: 14px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border: none; border-radius: 8px; font-size: 16px; cursor: pointer; margin-top: 10px; }
            .btn:hover { opacity: 0.9; }
            .message { margin-top: 15px; padding: 15px; border-radius: 8px; display: none; }
            .message.success { background: #d4edda; color: #155724; }
            .message.error { background: #f8d7da; color: #721c24; }
            .back-link { display: inline-block; margin-bottom: 20px; color: #667eea; text-decoration: none; }
            .card-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 20px; margin-top: 20px; }
            .card { background: white; border-radius: 12px; padding: 24px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
            .card h3 { font-size: 16px; color: #333; margin-bottom: 12px; display: flex; align-items: center; gap: 8px; }
            .card-btn { display: inline-block; padding: 12px 20px; background: linear-gradient(135deg, #3b82f6, #2563eb); color: white; border: none; border-radius: 8px; font-size: 14px; cursor: pointer; text-decoration: none; text-align: center; }
            .card-btn:hover { opacity: 0.9; }
            .card-btn.green { background: linear-gradient(135deg, #10b981, #059669); }
            .card-btn.red { background: linear-gradient(135deg, #ef4444, #dc2626); }
            .card-btn.pink { background: linear-gradient(135deg, #ec4899, #be185d); }
            .card-btn.orange { background: linear-gradient(135deg, #f59e0b, #d97706); }
        </style>
    </head>
    <body>
        <div class="header">
            <h1>📝 线索录入 - 管理后台</h1>
            <div class="nav">
                <a href="/">返回看板</a>
                <a href="/">看板首页</a>
            </div>
        </div>
        <div class="container" style="margin-bottom:20px">
            <a href="/" class="back-link">← 返回看板</a>
            <h2>录入新线索</h2>
            <form class="form" id="leadForm">
                <div class="form-group">
                    <label>手机号 *</label>
                    <input type="text" id="phone" required placeholder="请输入手机号">
                </div>
                <div class="form-group">
                    <label>线索平台 *</label>
                    <select id="platform" required>
                        <option value="">请选择</option>
                        <option value="400线索">400线索</option>
                        <option value="小红书">小红书</option>
                        <option value="转介绍">转介绍</option>
                        <option value="豆包">豆包</option>
                        <option value="其他">其他</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>录入日期 *</label>
                    <input type="date" id="entry_date" required value="''' + datetime.now().strftime('%Y-%m-%d') + '''">
                </div>
                <div class="form-group">
                    <label>分配给 *</label>
                    <select id="agent" required>
                        <option value="">请选择招商员</option>
                        ''' + '\n'.join([f'<option value="{a}">{a}</option>' for a in agents]) + '''
                    </select>
                </div>
                <button type="submit" class="btn">录入线索</button>
            </form>
            <div class="message" id="message"></div>
        </div>

        <!-- 快速功能入口 -->
        <div class="container">
            <h2>⚡ 快速功能</h2>
            <div class="card-grid">
                <div class="card">
                    <h3>💰 录入成本</h3>
                    <p style="color:#666;font-size:13px;margin-bottom:12px">录入抖音/小红书每日营销成本</p>
                    <button class="card-btn orange" onclick="openCostModal()">打开成本录入</button>
                </div>
                <div class="card">
                    <h3>📥 导入招商表</h3>
                    <p style="color:#666;font-size:13px;margin-bottom:12px">批量导入招商线索管理表</p>
                    <button class="card-btn blue" onclick="openZsImportModal()">打开导入窗口</button>
                </div>
                <div class="card">
                    <h3>🔥 导入抖音客资</h3>
                    <p style="color:#666;font-size:13px;margin-bottom:12px">批量导入抖音渠道线索</p>
                    <button class="card-btn red" onclick="openDyImportModal()">打开导入窗口</button>
                </div>
                <div class="card">
                    <h3>💖 导入小红书</h3>
                    <p style="color:#666;font-size:13px;margin-bottom:12px">批量导入小红书渠道线索</p>
                    <button class="card-btn pink" onclick="openXhsImportModal()">打开导入窗口</button>
                </div>
            </div>
        </div>

        <!-- 成本录入弹窗 -->
        <div id="costModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center">
            <div style="background:white;border-radius:12px;padding:30px;width:580px;max-width:95%;max-height:90vh;overflow-y:auto">
                <h3 style="margin-bottom:20px">💰 录入/管理营销成本</h3>

                <!-- 每日总消耗录入 -->
                <div style="background:#eff6ff;border-radius:10px;padding:16px;margin-bottom:16px">
                    <div style="font-weight:700;color:#1d4ed8;margin-bottom:12px">💰 录入每日总消耗</div>
                    <form id="costForm">
                        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:12px">
                            <div class="form-group">
                                <label style="font-size:13px;color:#666">日期</label>
                                <input type="date" id="costDate" required value="''' + datetime.now().strftime('%Y-%m-%d') + '''" style="padding:8px;border:1px solid #e0e0e0;border-radius:6px">
                            </div>
                            <div class="form-group">
                                <label style="font-size:13px;color:#666">平台</label>
                                <select id="costPlatform" required style="padding:8px;border:1px solid #e0e0e0;border-radius:6px">
                                    <option value="抖音">抖音</option>
                                    <option value="小红书">小红书</option>
                                </select>
                            </div>
                            <div class="form-group">
                                <label style="font-size:13px;color:#666">总消耗（元）</label>
                                <input type="number" id="costAmount" required placeholder="输入金额" step="0.01" style="padding:8px;border:1px solid #e0e0e0;border-radius:6px">
                            </div>
                        </div>
                        <button type="submit" class="card-btn orange" style="width:100%">确认录入</button>
                    </form>
                </div>

                <!-- 单条线索成本录入 -->
                <div style="background:#fff7ed;border-radius:10px;padding:16px;margin-bottom:16px">
                    <div style="font-weight:700;color:#c2410c;margin-bottom:12px">📊 录入单条线索成本</div>
                    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:12px">
                        <div class="form-group">
                            <label style="font-size:13px;color:#666">日期</label>
                            <input type="date" id="costUnitDate" value="''' + datetime.now().strftime('%Y-%m-%d') + '''" style="padding:8px;border:1px solid #e0e0e0;border-radius:6px">
                        </div>
                        <div class="form-group">
                            <label style="font-size:13px;color:#666">平台</label>
                            <select id="costUnitPlatform" style="padding:8px;border:1px solid #e0e0e0;border-radius:6px">
                                <option value="抖音">抖音</option>
                                <option value="小红书">小红书</option>
                            </select>
                        </div>
                        <div class="form-group">
                            <label style="font-size:13px;color:#666">单条成本（元/条）</label>
                            <input type="number" id="costUnit" placeholder="手动指定" step="0.01" style="padding:8px;border:1px solid #e0e0e0;border-radius:6px">
                        </div>
                    </div>
                    <button type="button" onclick="submitUnitCost()" class="card-btn" style="width:100%;background:#c2410c">录入单条成本</button>
                </div>

                <!-- 历史记录 -->
                <div style="margin-top:16px">
                    <div style="font-weight:700;color:#333;margin-bottom:10px">🗑 删除历史记录</div>
                    <div id="costHistory" style="border:1px solid #e0e0e0;border-radius:8px;max-height:200px;overflow-y:auto">
                        <div style="color:#999;text-align:center;padding:20px">加载中...</div>
                    </div>
                </div>

                <div class="message" id="costMessage" style="margin-top:15px"></div>
                <div style="display:flex;justify-content:flex-end;margin-top:16px">
                    <button type="button" class="card-btn" style="background:#6b7280" onclick="closeCostModal()">关闭</button>
                </div>
            </div>
        </div>

        <!-- 招商表导入弹窗 -->
        <div id="zsImportModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center">
            <div style="background:white;border-radius:12px;padding:30px;width:500px;max-width:90%">
                <h3 style="margin-bottom:20px">📥 导入招商表</h3>
                <div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;padding:12px;margin-bottom:15px;font-size:13px;color:#0369a1">
                    支持增量更新：已存在手机号的线索会更新其他信息，抖音/小红书平台的入库日期不会修改。
                </div>
                <form id="zsImportForm">
                    <div class="form-group" style="margin-bottom:15px">
                        <label>选择 Excel 文件</label>
                        <input type="file" id="zsExcelFile" accept=".xlsx,.xls" required style="padding:8px;border:2px solid #e0e0e0;border-radius:8px;width:100%">
                    </div>
                    <div style="display:flex;gap:10px;margin-top:20px">
                        <button type="submit" class="card-btn blue" style="flex:1">导入</button>
                        <button type="button" class="card-btn" style="flex:1;background:#6b7280" onclick="closeZsImportModal()">取消</button>
                    </div>
                </form>
                <div class="message" id="zsImportMessage" style="margin-top:15px"></div>
                <div id="zsImportResult" style="margin-top:15px"></div>
            </div>
        </div>

        <!-- 抖音导入弹窗 -->
        <div id="dyImportModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center">
            <div style="background:white;border-radius:12px;padding:30px;width:500px;max-width:90%">
                <h3 style="margin-bottom:20px">🔥 导入抖音客资</h3>
                <div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;padding:12px;margin-bottom:15px;font-size:13px;color:#0369a1">
                    只导入新线索：已存在手机号的线索会跳过，不更新已有数据。
                </div>
                <form id="dyImportForm2">
                    <div class="form-group" style="margin-bottom:15px">
                        <label>选择 Excel 文件</label>
                        <input type="file" id="dyExcelFile2" accept=".xlsx,.xls" required style="padding:8px;border:2px solid #e0e0e0;border-radius:8px;width:100%">
                    </div>
                    <div style="display:flex;gap:10px;margin-top:20px">
                        <button type="submit" class="card-btn red" style="flex:1">导入</button>
                        <button type="button" class="card-btn" style="flex:1;background:#6b7280" onclick="closeDyImportModal()">取消</button>
                    </div>
                </form>
                <div class="message" id="dyImportMessage2" style="margin-top:15px"></div>
                <div id="dyImportResult2" style="margin-top:15px"></div>
            </div>
        </div>

        <!-- 小红书导入弹窗 -->
        <div id="xhsImportModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center">
            <div style="background:white;border-radius:12px;padding:30px;width:500px;max-width:90%">
                <h3 style="margin-bottom:20px">💖 导入小红书</h3>
                <div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;padding:12px;margin-bottom:15px;font-size:13px;color:#0369a1">
                    只导入新线索：已存在手机号的线索会跳过，不更新已有数据。
                </div>
                <form id="xhsImportForm">
                    <div class="form-group" style="margin-bottom:15px">
                        <label>选择 Excel 文件</label>
                        <input type="file" id="xhsExcelFile" accept=".xlsx,.xls" required style="padding:8px;border:2px solid #e0e0e0;border-radius:8px;width:100%">
                    </div>
                    <div style="display:flex;gap:10px;margin-top:20px">
                        <button type="submit" class="card-btn pink" style="flex:1">导入</button>
                        <button type="button" class="card-btn" style="flex:1;background:#6b7280" onclick="closeXhsImportModal()">取消</button>
                    </div>
                </form>
                <div class="message" id="xhsImportMessage" style="margin-top:15px"></div>
                <div id="xhsImportResult" style="margin-top:15px"></div>
            </div>
        </div>

        <script>
        function openCostModal() {
            document.getElementById('costModal').style.display = 'flex';
            loadCostHistory();
        }
        function closeCostModal() { document.getElementById('costModal').style.display = 'none'; }
        function openZsImportModal() { document.getElementById('zsImportModal').style.display = 'flex'; }
        function closeZsImportModal() { document.getElementById('zsImportModal').style.display = 'none'; }
        function openDyImportModal() { document.getElementById('dyImportModal').style.display = 'flex'; }
        function closeDyImportModal() { document.getElementById('dyImportModal').style.display = 'none'; }
        function openXhsImportModal() { document.getElementById('xhsImportModal').style.display = 'flex'; }
        function closeXhsImportModal() { document.getElementById('xhsImportModal').style.display = 'none'; }

        // 加载成本历史记录
        async function loadCostHistory() {
            try {
                const res = await fetch('/api/cost');
                const data = await res.json();
                const container = document.getElementById('costHistory');
                const records = data.cost_data || [];

                if (records.length === 0) {
                    container.innerHTML = '<div style="color:#999;text-align:center;padding:20px">暂无成本记录</div>';
                    return;
                }

                let html = '<table style="width:100%;border-collapse:collapse;font-size:13px">';
                html += '<tr style="background:#f5f5f5"><th style="padding:10px;text-align:left;border-bottom:1px solid #e0e0e0">日期</th><th style="padding:10px;text-align:left;border-bottom:1px solid #e0e0e0">平台</th><th style="padding:10px;text-align:right;border-bottom:1px solid #e0e0e0">总消耗</th><th style="padding:10px;text-align:right;border-bottom:1px solid #e0e0e0">单条成本</th><th style="padding:10px;text-align:center;border-bottom:1px solid #e0e0e0">操作</th></tr>';
                records.forEach(r => {
                    html += '<tr>';
                    html += '<td style="padding:10px;border-bottom:1px solid #e0e0e0">' + r.date + '</td>';
                    html += '<td style="padding:10px;border-bottom:1px solid #e0e0e0">' + r.platform + '</td>';
                    html += '<td style="padding:10px;border-bottom:1px solid #e0e0e0;text-align:right">' + (r.amount > 0 ? '¥' + r.amount.toFixed(2) : '-') + '</td>';
                    html += '<td style="padding:10px;border-bottom:1px solid #e0e0e0;text-align:right">' + (r.unit_cost > 0 ? '¥' + r.unit_cost.toFixed(2) : '-') + '</td>';
                    html += "<td style='padding:10px;border-bottom:1px solid #e0e0e0;text-align:center'><button onclick=\"deleteCost('" + r.date + "', '" + r.platform + "')\" style='padding:4px 10px;font-size:12px;color:#ef4444;border:1px solid #ef4444;background:#fff;border-radius:4px;cursor:pointer'>删除</button></td>";
                    html += '</tr>';
                });
                html += '</table>';
                container.innerHTML = html;
            } catch (err) {
                document.getElementById('costHistory').innerHTML = '<div style="color:#ef4444;text-align:center;padding:20px">加载失败</div>';
            }
        }

        // 删除成本记录
        async function deleteCost(date, platform) {
            if (!confirm('确定删除 ' + date + ' ' + platform + ' 的成本记录？')) return;
            try {
                const res = await fetch('/api/cost/delete', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({cost_date: date, platform: platform})
                });
                const data = await res.json();
                const msg = document.getElementById('costMessage');
                msg.style.display = 'block';
                msg.className = 'message ' + (data.success ? 'success' : 'error');
                msg.textContent = data.message;
                if (data.success) loadCostHistory();
            } catch (err) {
                const msg = document.getElementById('costMessage');
                msg.style.display = 'block';
                msg.className = 'message error';
                msg.textContent = '删除失败';
            }
        }

        // 单条线索成本录入
        async function submitUnitCost() {
            const date = document.getElementById('costUnitDate').value;
            const platform = document.getElementById('costUnitPlatform').value;
            const unitCost = parseFloat(document.getElementById('costUnit').value);
            if (!date) { alert('请选择日期'); return; }
            if (isNaN(unitCost) || unitCost < 0) { alert('请输入有效的单条成本'); return; }

            try {
                const res = await fetch('/api/cost/add', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({cost_date: date, platform: platform, amount: 0, unit_cost: unitCost})
                });
                const data = await res.json();
                const msg = document.getElementById('costMessage');
                msg.style.display = 'block';
                msg.className = 'message ' + (data.success ? 'success' : 'error');
                msg.textContent = data.success ? '单条成本录入成功' : data.message;
                if (data.success) {
                    document.getElementById('costUnit').value = '';
                    loadCostHistory();
                }
            } catch (err) {
                const msg = document.getElementById('costMessage');
                msg.style.display = 'block';
                msg.className = 'message error';
                msg.textContent = '录入失败';
            }
        }

        // 每日总消耗录入提交
        document.getElementById('costForm').onsubmit = async (e) => {
            e.preventDefault();
            const res = await fetch('/api/cost/add', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({
                    cost_date: document.getElementById('costDate').value,
                    platform: document.getElementById('costPlatform').value,
                    amount: parseFloat(document.getElementById('costAmount').value)
                })
            });
            const data = await res.json();
            const msg = document.getElementById('costMessage');
            msg.style.display = 'block';
            msg.className = 'message ' + (data.success ? 'success' : 'error');
            msg.textContent = data.message;
            if (data.success) loadCostHistory();
        };

        // 招商表导入提交
        document.getElementById('zsImportForm').onsubmit = async (e) => {
            e.preventDefault();
            const fileInput = document.getElementById('zsExcelFile');
            const msgBox = document.getElementById('zsImportMessage');
            const resultBox = document.getElementById('zsImportResult');
            if (!fileInput.files[0]) { alert('请选择文件'); return; }
            msgBox.style.display = 'none'; resultBox.style.display = 'none';
            const formData = new FormData();
            formData.append('file', fileInput.files[0]);
            formData.append('type', 'zhaoshang');
            try {
                const res = await fetch('/api/leads/import', { method: 'POST', body: formData });
                const data = await res.json();
                msgBox.style.display = 'block';
                msgBox.className = 'message ' + (data.success ? 'success' : 'error');
                msgBox.textContent = data.message;
                if (data.success) {
                    let html = '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-top:10px">';
                    html += '<div style="background:#d4edda;border-radius:10px;padding:15px;text-align:center"><div style="font-size:24px;font-weight:700;color:#155724">' + (data.added || 0) + '</div><div style="font-size:13px;color:#155724">新增</div></div>';
                    html += '<div style="background:#cff4fc;border-radius:10px;padding:15px;text-align:center"><div style="font-size:24px;font-weight:700;color:#0c5460">' + (data.updated || 0) + '</div><div style="font-size:13px;color:#0c5460">更新</div></div>';
                    html += '<div style="background:#fff3cd;border-radius:10px;padding:15px;text-align:center"><div style="font-size:24px;font-weight:700;color:#856404">' + (data.skipped || 0) + '</div><div style="font-size:13px;color:#856404">跳过</div></div>';
                    html += '</div>';
                    resultBox.innerHTML = html; resultBox.style.display = 'block';
                }
            } catch(err) { msgBox.style.display = 'block'; msgBox.className = 'message error'; msgBox.textContent = '网络错误'; }
        };

        // 抖音导入提交
        document.getElementById('dyImportForm2').onsubmit = async (e) => {
            e.preventDefault();
            const fileInput = document.getElementById('dyExcelFile2');
            const msgBox = document.getElementById('dyImportMessage2');
            const resultBox = document.getElementById('dyImportResult2');
            if (!fileInput.files[0]) { alert('请选择文件'); return; }
            msgBox.style.display = 'none'; resultBox.style.display = 'none';
            const formData = new FormData();
            formData.append('file', fileInput.files[0]);
            try {
                const res = await fetch('/api/leads/import-douyin', { method: 'POST', body: formData });
                const data = await res.json();
                msgBox.style.display = 'block';
                msgBox.className = 'message ' + (data.success ? 'success' : 'error');
                msgBox.textContent = data.message;
                if (data.success) {
                    let html = '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-top:10px">';
                    html += '<div style="background:#d4edda;border-radius:10px;padding:15px;text-align:center"><div style="font-size:24px;font-weight:700;color:#155724">' + (data.added || 0) + '</div><div style="font-size:13px;color:#155724">新增</div></div>';
                    html += '<div style="background:#fff3cd;border-radius:10px;padding:15px;text-align:center"><div style="font-size:24px;font-weight:700;color:#856404">' + (data.skipped || 0) + '</div><div style="font-size:13px;color:#856404">重复跳过</div></div>';
                    html += '<div style="background:#f8d7da;border-radius:10px;padding:15px;text-align:center"><div style="font-size:24px;font-weight:700;color:#721c24">' + (data.bad || 0) + '</div><div style="font-size:13px;color:#721c24">无法识别</div></div>';
                    html += '</div>';
                    resultBox.innerHTML = html; resultBox.style.display = 'block';
                }
            } catch(err) { msgBox.style.display = 'block'; msgBox.className = 'message error'; msgBox.textContent = '网络错误'; }
        };

        // 小红书导入提交
        document.getElementById('xhsImportForm').onsubmit = async (e) => {
            e.preventDefault();
            const fileInput = document.getElementById('xhsExcelFile');
            const msgBox = document.getElementById('xhsImportMessage');
            const resultBox = document.getElementById('xhsImportResult');
            if (!fileInput.files[0]) { alert('请选择文件'); return; }
            msgBox.style.display = 'none'; resultBox.style.display = 'none';
            const formData = new FormData();
            formData.append('file', fileInput.files[0]);
            formData.append('type', 'xiaohongshu');
            try {
                const res = await fetch('/api/leads/import', { method: 'POST', body: formData });
                const data = await res.json();
                msgBox.style.display = 'block';
                msgBox.className = 'message ' + (data.success ? 'success' : 'error');
                msgBox.textContent = data.message;
                if (data.success) {
                    let html = '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-top:10px">';
                    html += '<div style="background:#d4edda;border-radius:10px;padding:15px;text-align:center"><div style="font-size:24px;font-weight:700;color:#155724">' + (data.added || 0) + '</div><div style="font-size:13px;color:#155724">新增</div></div>';
                    html += '<div style="background:#fff3cd;border-radius:10px;padding:15px;text-align:center"><div style="font-size:24px;font-weight:700;color:#856404">' + (data.skipped || 0) + '</div><div style="font-size:13px;color:#856404">重复跳过</div></div>';
                    html += '<div style="background:#f8d7da;border-radius:10px;padding:15px;text-align:center"><div style="font-size:24px;font-weight:700;color:#721c24">' + (data.bad || 0) + '</div><div style="font-size:13px;color:#721c24">无法识别</div></div>';
                    html += '</div>';
                    resultBox.innerHTML = html; resultBox.style.display = 'block';
                }
            } catch(err) { msgBox.style.display = 'block'; msgBox.className = 'message error'; msgBox.textContent = '网络错误'; }
        };
        </script>

        <div class="container" style="margin-top:20px">
            <h2>📊 抖音客资批量导入</h2>
            <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:10px;padding:16px 20px;margin-bottom:16px;font-size:13px;color:#166534;line-height:1.7">
                <div style="font-weight:700;margin-bottom:4px">&#128161; 导入说明</div>
                <div>• 支持 .xlsx / .xls 格式</div>
                <div>• 自动识别列：线索创建时间 → 入库日期 | 姓名 → 姓名 | 手机号 → 手机号 | 跟进员工 → 招商员 | 所在城市 → 城市</div>
                <div>• <b>手机号重复</b> → 直接跳过，不更新不插入</div>
                <div>• 其他字段信息不填入系统</div>
            </div>
            <form id="dyImportForm">
                <div class="form-group">
                    <label>选择抖音客资 Excel 文件 <span style="color:#999;font-weight:400">（.xlsx / .xls）</span></label>
                    <input type="file" id="dyExcelFile" accept=".xlsx,.xls" required style="padding:8px;border:2px solid #e0e0e0;border-radius:8px">
                </div>
                <button type="submit" class="btn" id="dyImportBtn" style="background:linear-gradient(135deg,#10b981,#059669)">导入表格</button>
            </form>
            <div class="message" id="dyImportMessage"></div>
            <div id="dyImportResult" style="margin-top:15px;display:none"></div>
        </div>

        <script>
            document.getElementById('leadForm').onsubmit = async (e) => {
                e.preventDefault();
                const res = await fetch('/api/leads/add', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        phone: document.getElementById('phone').value,
                        platform: document.getElementById('platform').value,
                        entry_date: document.getElementById('entry_date').value,
                        agent: document.getElementById('agent').value
                    })
                });
                const data = await res.json();
                const msg = document.getElementById('message');
                msg.style.display = 'block';
                msg.className = 'message ' + (data.success ? 'success' : 'error');
                msg.textContent = data.message;
                if (data.success) {
                    document.getElementById('phone').value = '';
                }
            };

            document.getElementById('dyImportForm').onsubmit = async (e) => {
                e.preventDefault();
                const fileInput = document.getElementById('dyExcelFile');
                const btn = document.getElementById('dyImportBtn');
                const msgBox = document.getElementById('dyImportMessage');
                const resultBox = document.getElementById('dyImportResult');

                if (!fileInput.files[0]) {
                    msgBox.style.display = 'block';
                    msgBox.className = 'message error';
                    msgBox.textContent = '请选择Excel文件';
                    return;
                }

                btn.disabled = true;
                btn.textContent = '正在导入...';
                msgBox.style.display = 'none';
                resultBox.style.display = 'none';

                const formData = new FormData();
                formData.append('file', fileInput.files[0]);

                try {
                    const res = await fetch('/api/leads/import-douyin', { method: 'POST', body: formData });
                    const data = await res.json();

                    msgBox.style.display = 'block';
                    msgBox.className = 'message ' + (data.success ? 'success' : 'error');
                    msgBox.textContent = data.message;

                    if (data.success) {
                        fileInput.value = '';
                        let html = '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-top:10px">';
                        html += '<div style="background:#d4edda;border-radius:10px;padding:15px;text-align:center"><div style="font-size:24px;font-weight:700;color:#155724">' + (data.added || 0) + '</div><div style="font-size:13px;color:#155724">新增</div></div>';
                        html += '<div style="background:#fff3cd;border-radius:10px;padding:15px;text-align:center"><div style="font-size:24px;font-weight:700;color:#856404">' + (data.skipped || 0) + '</div><div style="font-size:13px;color:#856404">重复跳过</div></div>';
                        html += '<div style="background:#f8d7da;border-radius:10px;padding:15px;text-align:center"><div style="font-size:24px;font-weight:700;color:#721c24">' + (data.bad || 0) + '</div><div style="font-size:13px;color:#721c24">无法识别</div></div>';
                        html += '</div>';
                        resultBox.innerHTML = html;
                        resultBox.style.display = 'block';
                    }
                } catch(err) {
                    msgBox.style.display = 'block';
                    msgBox.className = 'message error';
                    msgBox.textContent = '网络错误: ' + err;
                } finally {
                    btn.disabled = false;
                    btn.textContent = '导入表格';
                }
            };
        </script>
    </body>
    </html>
    '''
    return html

# ─────────────────────────────────────────────
# 主页
# ─────────────────────────────────────────────
@app.route('/', methods=['GET'])
def index():
    user = session.get('user')
    if not user:
        return login_page()
    
    user_name = user['name']
    role_text = '管理员' if user['role'] == 'admin' else '招商员'
    
    html = '''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>线索看板</title>
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            body { font-family: -apple-system, BlinkMacSystemFont, sans-serif; background: #f5f7fa; }
            .header { 
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white; padding: 20px 40px; display: flex; justify-content: space-between; align-items: center;
            }
            .header h1 { font-size: 24px; }
            .user-info { display: flex; align-items: center; gap: 20px; }
            .nav-links { display: flex; gap: 15px; }
            .nav-links a { color: white; text-decoration: none; padding: 8px 16px; background: rgba(255,255,255,0.2); border-radius: 8px; font-size: 14px; }
            .nav-links a:hover { background: rgba(255,255,255,0.3); }
            .badge { background: #ef4444; color: white; padding: 2px 8px; border-radius: 10px; font-size: 12px; margin-left: 5px; }
            .logout-btn { 
                background: rgba(255,255,255,0.2); color: white; border: 1px solid rgba(255,255,255,0.3);
                padding: 8px 20px; border-radius: 20px; cursor: pointer; font-size: 14px;
            }
            .logout-btn:hover { background: rgba(255,255,255,0.3); }
            .container { padding: 20px 40px; }
            iframe { width: 100%; height: calc(100vh - 100px); border: none; border-radius: 12px; background: white; }
        </style>
    </head>
    <body>
        <div class="header">
            <h1>📊 线索看板</h1>
            <div class="user-info">
                <span>👤 ''' + user_name + ''' (''' + role_text + ''')</span>
                <div class="nav-links">
                    ''' + ('<a href="/admin">📝 录入线索</a>' if user['role'] == 'admin' else '') + '''
                </div>
                <button class="logout-btn" onclick="logout()">退出登录</button>
            </div>
        </div>
        <div class="container">
            <iframe src="/kanban_content"></iframe>
        </div>
        <script>
            async function logout() {
                await fetch('/api/logout', {method: 'POST'});
                window.location.href = '/';
            }
        </script>
    </body>
    </html>
    '''
    return html

# ─────────────────────────────────────────────
# 登录页面
# ─────────────────────────────────────────────
@app.route('/login', methods=['GET'])
def login_page():
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>线索看板 - 登录</title>
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            body { 
                font-family: -apple-system, BlinkMacSystemFont, sans-serif;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                min-height: 100vh;
                display: flex;
                align-items: center;
                justify-content: center;
            }
            .login-box {
                background: white;
                padding: 40px;
                border-radius: 16px;
                box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                width: 380px;
            }
            h1 { text-align: center; color: #333; margin-bottom: 30px; font-size: 24px; }
            .input-group { margin-bottom: 20px; }
            label { display: block; margin-bottom: 8px; color: #666; font-size: 14px; }
            input { 
                width: 100%; padding: 12px 16px; border: 2px solid #e0e0e0;
                border-radius: 8px; font-size: 16px;
            }
            input:focus { outline: none; border-color: #667eea; }
            button { 
                width: 100%; padding: 14px;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white; border: none; border-radius: 8px;
                font-size: 16px; font-weight: 600; cursor: pointer;
            }
            button:hover { transform: translateY(-2px); }
            .error { color: #e74c3c; text-align: center; margin-top: 16px; font-size: 14px; }
            .demo { margin-top: 20px; padding-top: 20px; border-top: 1px solid #eee; font-size: 12px; color: #999; }
            .demo h3 { font-size: 13px; margin-bottom: 10px; color: #666; }
            .demo p { margin: 5px 0; }
            code { background: #f5f5f5; padding: 2px 6px; border-radius: 4px; }
        </style>
    </head>
    <body>
        <div class="login-box">
            <h1>🔐 线索看板登录</h1>
            <form id="loginForm">
                <div class="input-group">
                    <label>用户名</label>
                    <input type="text" id="username" name="username" required>
                </div>
                <div class="input-group">
                    <label>密码</label>
                    <input type="password" id="password" name="password" required>
                </div>
                <button type="submit">登录</button>
                <div class="error" id="error"></div>
            </form>

        </div>
        <script>
            document.getElementById('loginForm').onsubmit = async (e) => {
                e.preventDefault();
                const res = await fetch('/api/login', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        username: document.getElementById('username').value,
                        password: document.getElementById('password').value
                    })
                });
                const data = await res.json();
                if (data.success) {
                    window.location.href = '/';
                } else {
                    document.getElementById('error').textContent = data.message;
                }
            };
        </script>
    </body>
    </html>
    '''

# ─────────────────────────────────────────────
# 看板内容（注入过滤后的数据）
# ─────────────────────────────────────────────
@app.route('/kanban_content', methods=['GET'])
def kanban_content():
    user = session.get('user')
    if not user:
        return '<h1>请先登录</h1><a href="/">返回登录</a>'
    
    html_file = BASE_DIR / '线索看板.html'
    with open(html_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    records = load_data()
    new_leads = load_new_leads()
    
    # 根据用户角色过滤数据（管理员和游客看全部）
    if user['role'] == 'admin' or user['role'] == 'guest':
        filtered = records + new_leads
    else:
        agent_name = user['name']
        filtered = [r for r in records if r.get('所属招商', '') == agent_name 
                    or r.get('跟进员工', '') == agent_name]
        # 加上新录入的线索
        filtered += [r for r in new_leads if r.get('所属招商', '') == agent_name]
    
    # 统一按入库时间降序排列（最近日期在前）
    filtered.sort(key=lambda x: str(x.get('入库时间', '') or x.get('录入日期', '') or ''), reverse=True)
    
    # 替换原始数据
    pattern = r'window\.__ALL__\s*=\s*\[[\s\S]*?\];'
    replacement = 'window.__ALL__ = ' + json.dumps(filtered, ensure_ascii=False) + ';'
    content = re.sub(pattern, replacement, content, count=1)

    # 注入成本数据（从数据库实时读取，按平台区分）
    cost_data = load_cost_data()
    cost_by_plat = {'抖音': {}, '小红书': {}}
    all_days = set()
    for c in cost_data:
        d, plat, amt, uc = c['date'], c['platform'], c['amount'], c.get('unit_cost', 0)
        all_days.add(d)
        if plat not in cost_by_plat:
            cost_by_plat[plat] = {}
        cost_by_plat[plat][d] = {'spend': amt, 'unit_cost': uc if uc else 0, 'leads': 0}
    # 按平台计算单条成本（若数据库中未填写则自动计算）
    for plat in cost_by_plat:
        for d in cost_by_plat[plat]:
            day_leads = len([r for r in filtered if str(r.get('入库时间', '')).startswith(d) and r.get('平台') == plat])
            cost_by_plat[plat][d]['leads'] = day_leads
            if cost_by_plat[plat][d]['unit_cost'] <= 0 and day_leads > 0 and cost_by_plat[plat][d]['spend'] > 0:
                cost_by_plat[plat][d]['unit_cost'] = round(cost_by_plat[plat][d]['spend'] / day_leads, 2)
    cost_days = sorted(all_days)
    cost_script = 'window.__COST_BY_PLAT__ = ' + json.dumps(cost_by_plat, ensure_ascii=False) + ';\nwindow.__COST_DAYS__ = ' + json.dumps(cost_days, ensure_ascii=False) + ';'
    content = re.sub(r'window\.__COST_BY_PLAT__\s*=\s*\{[\s\S]*?\};\s*window\.__COST_DAYS__\s*=\s*\[[\s\S]*?\];', cost_script, content, count=1)
    # 兼容旧变量（HTML 中已改为 COST_BY_PLAT，此行为保险保留）
    content = content.replace('window.__COST__', 'window.__COST_BY_PLAT__')

    # 注入用户信息和未读提醒
    unread_count = 0
    unread_leads = []
    if user['role'] != 'admin':
        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()
        c.execute('SELECT * FROM new_leads WHERE agent = ? AND is_read = 0', (user['name'],))
        rows = c.fetchall()
        conn.close()
        unread_count = len(rows)
        for row in rows:
            unread_leads.append({'id': row[0], '手机号': row[1], '平台': row[2], '入库时间': row[10][:10] if row[10] else ''})
    
    # 注入用户信息和提醒
    user_script = '<script>window.CURRENT_USER = ' + json.dumps(user) + '; window.UNREAD_LEADS = ' + json.dumps(unread_leads) + '; window.UNREAD_COUNT = ' + str(unread_count) + ';</script>'
    content = content.replace('</head>', user_script + '</head>')
    
    # 添加提醒弹窗样式和逻辑
    if user['role'] != 'admin' and unread_count > 0:
        alert_html = '''
        <div id="leadAlert" style="position:fixed;top:70px;right:20px;background:linear-gradient(135deg,#f59e0b,#ef4444);color:white;padding:20px;border-radius:12px;box-shadow:0 4px 20px rgba(0,0,0,0.3);z-index:9999;max-width:350px;display:none;">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;">
                <strong style="font-size:16px;">🔔 新线索提醒</strong>
                <span id="alertClose" style="cursor:pointer;font-size:20px;padding:0 5px;">×</span>
            </div>
            <p style="margin-bottom:10px;">您有 <strong id="alertCount">''' + str(unread_count) + '''</strong> 条新线索待处理：</p>
            <div id="alertList"></div>
        </div>
        <script>
            (function() {
                var leads = window.UNREAD_LEADS || [];
                var count = window.UNREAD_COUNT || 0;
                if (count > 0) {
                    var alertDiv = document.getElementById('leadAlert');
                    var listDiv = document.getElementById('alertList');
                    leads.forEach(function(lead) {
                        listDiv.innerHTML += '<div style="background:rgba(255,255,255,0.2);padding:8px 10px;border-radius:6px;margin-top:8px;font-size:13px;">📱 ' + lead['手机号'] + '<br><span style="font-size:12px;">平台: ' + lead['平台'] + ' | ' + lead['入库时间'] + '</span></div>';
                    });
                    alertDiv.style.display = 'block';
                    document.getElementById('alertClose').onclick = function() {
                        alertDiv.style.display = 'none';
                        // 标记所有未读线索为已读
                        leads.forEach(function(lead) {
                            fetch('/api/leads/mark_read', {
                                method: 'POST',
                                headers: {'Content-Type': 'application/json'},
                                body: JSON.stringify({id: lead.id})
                            });
                        });
                    };
                }
            })();
        </script>
        '''
        content = content.replace('</body>', alert_html + '</body>')
    
    return content

# ─────────────────────────────────────────────
# 启动时修复平台分类
# ─────────────────────────────────────────────
def fix_platform_classification():
    """启动时将'抖音广告'改为'抖音'"""
    try:
        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM new_leads WHERE platform = '抖音广告'")
        count = c.fetchone()[0]
        if count > 0:
            c.execute("UPDATE new_leads SET platform = '抖音' WHERE platform = '抖音广告'")
            conn.commit()
            print(f"[启动修复] 已将 {count} 条'抖音广告'改为'抖音'")
        conn.close()
    except Exception as e:
        print(f"[启动修复] 修复平台分类失败: {e}")

if __name__ == '__main__':
    # 启动时自动修复平台分类
    fix_platform_classification()
    
    print("=" * 50)
    print("🔐 线索看板服务 v2 启动中...")
    print("📍 访问地址: http://localhost:5001")
    print("=" * 50)
    print("\n测试账号：")
    print("  管理员: admin / admin123")
    print("  招商员1: zhengjianjun / zjj001345")
    print("  招商员2: liurenjie / lrj001678")
    print()
    app.run(debug=False, port=5001, host='0.0.0.0')
