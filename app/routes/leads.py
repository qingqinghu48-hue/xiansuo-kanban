"""
线索路由蓝图
"""

import re
import sqlite3
from datetime import datetime

from flask import Blueprint, jsonify, request, session

from app.config import Config
from app.models.leads import load_data, load_new_leads

DB_FILE = Config.DB_FILE

leads_bp = Blueprint("leads", __name__)


@leads_bp.route("/api/leads", methods=["GET"])
def get_leads():
    user = session.get("user")
    if not user:
        return jsonify({"error": "请先登录"}), 401

    records = load_data()
    new_leads = load_new_leads()

    # 管理员和游客看全部
    if user["role"] == "admin" or user["role"] == "guest":
        all_records = records + new_leads
        return jsonify(
            {
                "records": all_records,
                "total": len(all_records),
                "new_leads_count": len(new_leads),
            }
        )

    # 招商员只看自己分配的
    agent_name = user["name"]
    filtered = [
        r
        for r in records
        if r.get("所属招商", "") == agent_name or r.get("跟进员工", "") == agent_name
    ]

    # 加上新录入的线索
    agent_new_leads = [r for r in new_leads if r.get("所属招商", "") == agent_name]
    all_filtered = filtered + agent_new_leads

    # 获取未读新线索数
    unread = len([r for r in agent_new_leads if r.get("是否已读", 1) == 0])

    return jsonify(
        {
            "records": all_filtered,
            "total": len(all_filtered),
            "role": "agent",
            "new_leads_count": len(agent_new_leads),
            "unread_count": unread,
        }
    )


@leads_bp.route("/api/leads/add", methods=["POST"])
def add_lead():
    user = session.get("user")
    if not user or user["role"] != "admin":
        return jsonify({"success": False, "message": "只有管理员可以录入线索"}), 401

    data = request.json
    phone = data.get("phone", "").strip()
    platform = data.get("platform", "").strip()
    agent = data.get("agent", "").strip()
    entry_date = data.get("entry_date", "").strip() or datetime.now().strftime("%Y-%m-%d")

    if not phone or not platform or not agent:
        return jsonify({"success": False, "message": "请填写完整信息"})

    # 检查是否已存在
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute("SELECT id FROM new_leads WHERE phone = ?", (phone,))
    if c.fetchone():
        conn.close()
        return jsonify({"success": False, "message": "该手机号已录入"})

    # 插入新线索
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute(
        """
        INSERT INTO new_leads (phone, platform, agent, entry_date, created_at)
        VALUES (?, ?, ?, ?, ?)
    """,
        (phone, platform, agent, entry_date, now),
    )
    conn.commit()
    conn.close()

    return jsonify({"success": True, "message": "线索录入成功"})


@leads_bp.route("/api/leads/update", methods=["POST"])
def update_lead():
    user = session.get("user")
    if not user:
        return jsonify({"success": False, "message": "请先登录"}), 401

    data = request.json
    phone = data.get("phone", "").strip()

    # 检查是否是该招商员的线索
    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute("SELECT id, agent FROM new_leads WHERE phone = ?", (phone,))
    row = c.fetchone()

    if not row:
        conn.close()
        return jsonify({"success": False, "message": "线索不存在"})

    lead_id, lead_agent = row

    # 管理员可以更新所有，招商员只能更新自己的
    if user["role"] != "admin" and lead_agent != user["name"]:
        conn.close()
        return jsonify({"success": False, "message": "无权修改此线索"})

    # 基础字段（所有用户可修改）
    name = data.get("name", "")
    city = data.get("city", "")
    validity = data.get("validity", "")
    region = data.get("region", "")
    can_wechat = data.get("can_wechat", "")
    remark = data.get("remark", "")
    entry_date = data.get("entry_date", "")
    contact_time = data.get("二次联系时间", "")
    contact_remark = data.get("二次联系备注", "")
    call_time = data.get("最近一次电联时间", "")
    visit_time = data.get("到访时间", "")
    sign_time = data.get("签约时间", "")
    platform = data.get("platform", "")
    xhs_account = data.get("xhs_account", "")
    lead_type = data.get("lead_type", "")

    # 管理员可修改平台和入库日期，普通招商员保留原值
    if user["role"] == "admin":
        update_sql = """
            UPDATE new_leads SET
                name = ?, city = ?, validity = ?, region = ?, can_wechat = ?, remark = ?,
                platform = ?, entry_date = ?, 二次联系时间 = ?, 二次联系备注 = ?, 最近一次电联时间 = ?, 到访时间 = ?, 签约时间 = ?,
                xhs_account = ?, lead_type = ?
            WHERE id = ?
        """
        update_params = (
            name,
            city,
            validity,
            region,
            can_wechat,
            remark,
            platform,
            entry_date,
            contact_time,
            contact_remark,
            call_time,
            visit_time,
            sign_time,
            xhs_account,
            lead_type,
            lead_id,
        )
    else:
        # 保留原平台和入库日期
        c.execute("SELECT platform, entry_date FROM new_leads WHERE id = ?", (lead_id,))
        orig = c.fetchone()
        orig_platform = orig[0] if orig else ""
        orig_entry_date = orig[1] if orig else ""
        _ = orig_platform  # unused but kept for fidelity
        _ = orig_entry_date  # unused but kept for fidelity
        update_sql = """
            UPDATE new_leads SET
                name = ?, city = ?, validity = ?, region = ?, can_wechat = ?, remark = ?,
                二次联系时间 = ?, 二次联系备注 = ?, 最近一次电联时间 = ?, 到访时间 = ?, 签约时间 = ?,
                xhs_account = ?, lead_type = ?
            WHERE id = ?
        """
        update_params = (
            name,
            city,
            validity,
            region,
            can_wechat,
            remark,
            contact_time,
            contact_remark,
            call_time,
            visit_time,
            sign_time,
            xhs_account,
            lead_type,
            lead_id,
        )

    c.execute(update_sql, update_params)
    conn.commit()
    conn.close()

    return jsonify({"success": True, "message": "更新成功"})


@leads_bp.route("/api/leads/delete", methods=["POST"])
def delete_lead():
    user = session.get("user")
    if not user:
        return jsonify({"success": False, "message": "请先登录"}), 401

    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({"success": False, "message": "请求数据为空"})
        lead_id = data.get("id")
        phone = str(data.get("phone", "")).strip()
        if not lead_id and not phone:
            return jsonify({"success": False, "message": "请提供id或手机号"})
    except Exception as e:
        return jsonify({"success": False, "message": "请求解析失败: " + str(e)})

    try:
        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()
        if lead_id:
            c.execute("SELECT id, agent FROM new_leads WHERE id = ?", (lead_id,))
        else:
            c.execute("SELECT id, agent FROM new_leads WHERE phone = ?", (phone,))
        row = c.fetchone()

        if not row:
            conn.close()
            return jsonify(
                {"success": False, "message": "线索不存在或无法删除（仅支持删除手动录入的线索）"}
            )

        lead_id, lead_agent = row

        # 管理员可以删除所有，招商员只能删除自己的
        if user["role"] != "admin" and lead_agent != user["name"]:
            conn.close()
            return jsonify({"success": False, "message": "无权删除此线索"})

        c.execute("DELETE FROM new_leads WHERE id = ?", (lead_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True, "message": "删除成功"})
    except Exception as e:
        return jsonify({"success": False, "message": "删除出错: " + str(e)})


@leads_bp.route("/api/leads/batch-delete", methods=["POST"])
def batch_delete_leads():
    user = session.get("user")
    if not user:
        return jsonify({"success": False, "message": "请先登录"}), 401

    try:
        data = request.get_json(force=True)
        phones = data.get("phones", [])
        if not phones or not isinstance(phones, list):
            return jsonify({"success": False, "message": "请提供要删除的手机号列表"})
    except Exception as e:
        return jsonify({"success": False, "message": "请求解析失败: " + str(e)})

    try:
        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()

        deleted = 0
        skipped = 0
        for phone in phones:
            phone = str(phone).strip()
            if not phone:
                skipped += 1
                continue

            c.execute("SELECT id, agent FROM new_leads WHERE phone = ?", (phone,))
            row = c.fetchone()
            if not row:
                skipped += 1
                continue

            lead_id, lead_agent = row
            # 权限检查
            if user["role"] != "admin" and lead_agent != user["name"]:
                skipped += 1
                continue

            c.execute("DELETE FROM new_leads WHERE id = ?", (lead_id,))
            deleted += 1

        conn.commit()
        conn.close()
        return jsonify(
            {
                "success": True,
                "message": f"批量删除完成：成功 {deleted} 条，跳过 {skipped} 条",
            }
        )
    except Exception as e:
        return jsonify({"success": False, "message": "删除出错: " + str(e)})


@leads_bp.route("/api/leads/mark_read", methods=["POST"])
def mark_lead_read():
    user = session.get("user")
    if not user:
        return jsonify({"success": False, "message": "请先登录"}), 401

    data = request.json
    lead_id = data.get("id")

    conn = sqlite3.connect(str(DB_FILE))
    c = conn.cursor()
    c.execute("UPDATE new_leads SET is_read = 1 WHERE id = ?", (lead_id,))
    conn.commit()
    conn.close()

    return jsonify({"success": True})


@leads_bp.route("/api/leads/import", methods=["POST"])
def import_leads():
    user = session.get("user")
    if not user or user["role"] != "admin":
        return (
            jsonify({"success": False, "message": "只有管理员可以导入线索"}),
            401,
        )

    if "file" not in request.files:
        return jsonify({"success": False, "message": "请选择文件"})

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"success": False, "message": "请选择文件"})

    try:
        import pandas as pd
        from io import BytesIO

        file_bytes = file.read()
        filename = file.filename
        filename_lower = filename.lower()

        # ── 1) 读取Excel ──
        bio = BytesIO(file_bytes)
        df = None
        try:
            if filename_lower.endswith(".xls"):
                df = pd.read_excel(bio, engine="xlrd")
            else:
                df = pd.read_excel(bio, engine="openpyxl")
        except Exception as read_err:
            err_msg = str(read_err)
            if "not supported" in err_msg.lower():
                err_msg += "（服务器缺少openpyxl库，请联系管理员安装: pip install openpyxl）"
            return jsonify({"success": False, "message": f"读取Excel失败: {err_msg}"})

        # 兼容 pandas 1.1.5：直接读取异常时回退到 openpyxl 手动构造
        if not filename_lower.endswith(".xls") and (
            df is None or len(df) == 0 or len(df.columns) <= 1
        ):
            try:
                bio.seek(0)
                from openpyxl import load_workbook

                wb = load_workbook(bio, data_only=True)
                ws = wb.active
                data = []
                headers = None
                for row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(c).strip() if c is not None else "" for c in row]
                    else:
                        data.append(row)
                df = pd.DataFrame(data, columns=headers)
                if df is not None and len(df) > 0:
                    df = df.dropna(how="all")
            except Exception:
                pass

        if df is None or len(df) == 0:
            return jsonify({"success": False, "message": "Excel 文件为空（0行数据）"})

        cols = [str(c).strip() for c in df.columns]
        print(f"[导入] 文件={filename}, 行数={len(df)}, 列名={cols}")

        # ── 2) 智能识别列 ──
        def find_col(keywords):
            for kw in keywords:
                kw_lower = kw.lower()
                for col in cols:
                    if kw_lower in col.lower():
                        return col
            return None

        phone_col = find_col(
            ["手机号", "手机号码", "电话", "联系电话", "手机", "phone", "客户电话"]
        )
        weixin_col = find_col(["微信号", "微信", "微信账号"])
        platform_col = find_col(["平台", "来源", "渠道", "来源平台", "线索来源"])
        agent_col = find_col(
            ["所属招商", "跟进员工", "负责人", "招商员", "员工", "分配"]
        )
        date_col = find_col(
            ["入库日期", "录入日期", "日期", "入库时间", "录入时间", "线索生成时间"]
        )
        # 小红书客资专用列
        xhs_account_col = find_col(["归属账号", "小红书账号", "账号"])
        lead_type_col = find_col(["流量类型", "线索类型", "类型"])
        name_col = find_col(["姓名", "名字", "客户姓名", "联系人"])
        city_col = find_col(["城市", "省份", "地区", "所在城市", "省"])
        region_col = find_col(["所属大区", "大区", "区域"])
        validity_col = find_col(["线索有效性", "有效性", "客户类型", "等级"])
        wechat_col = find_col(["是否能加上微信", "能否加微", "加微信", "微信"])
        remark_col = find_col(["备注", "说明", "备注信息", "客户情况备注"])
        follow_time_col = find_col(["二次联系时间", "跟进时间", "下次联系时间"])
        follow_note_col = find_col(["二次联系备注", "跟进备注", "联系备注"])
        call_time_col = find_col(["最近一次电联时间", "电联时间", "最近联系时间"])
        visit_time_col = find_col(["到访时间", "到店时间", "来访时间"])
        sign_time_col = find_col(["签约时间", "成交时间", "签约日期"])

        if not phone_col:
            return jsonify(
                {"success": False, "message": f"无法识别手机号列。当前列名: {cols}"}
            )

        print(
            f"[导入] 列映射: 手机={phone_col}, 平台={platform_col}, 招商={agent_col}, 日期={date_col}"
        )

        # ── 3) 文件类型判断 ──
        req_type = request.form.get("type", "")
        is_zhaoshang = "招商" in filename or req_type == "zhaoshang"
        is_douyin_channel = req_type == "douyin"
        is_xhs_channel = req_type == "xiaohongshu"
        is_douyin_kezi = "客资" in filename or (
            "抖音" in filename and not is_zhaoshang
        )

        # ── 4) 解析所有行 ──
        def get_val(row, col):
            if not col:
                return ""
            v = row.get(col, "")
            if pd.isna(v):
                return ""
            s = str(v).strip()
            return s if s.lower() != "nan" else ""

        def parse_date(row, col):
            if not col:
                return ""
            v = row.get(col)
            if pd.isna(v):
                return ""
            try:
                return pd.to_datetime(v).strftime("%Y-%m-%d")
            except (ValueError, TypeError):
                pass
            # 兼容中文日期格式：2026年4月14日
            s = str(v).strip()
            m = re.match(r"(\d{4})年(\d{1,2})月(\d{1,2})日", s)
            if m:
                return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
            return ""

        parsed = []
        bad_rows = []
        for idx, row in df.iterrows():
            # 小红书客资导入：优先手机号，没有则取微信号
            raw = row.get(phone_col, "") if phone_col else ""
            weixin_raw = row.get(weixin_col, "") if weixin_col else ""

            phone = ""
            if pd.notna(raw) and str(raw).strip() and str(raw).lower() != "nan":
                s = str(raw).strip()
                digits = "".join(filter(str.isdigit, s))
                if len(digits) >= 7:
                    phone = digits
                elif len(s) >= 5:  # 可能是微信号
                    phone = s
            elif (
                pd.notna(weixin_raw)
                and str(weixin_raw).strip()
                and str(weixin_raw).lower() != "nan"
            ):
                phone = str(weixin_raw).strip()

            if not phone:
                bad_rows.append({"row": int(idx) + 2, "raw": str(raw), "reason": "联系方式为空"})
                continue

            # 入库日期：以Excel表格为准
            entry_date = parse_date(row, date_col)
            if not entry_date:
                for fc in ["日期", "时间", "创建日期", "添加日期"]:
                    col = find_col([fc])
                    if col:
                        entry_date = parse_date(row, col)
                        if entry_date:
                            break
            if not entry_date:
                entry_date = datetime.now().strftime("%Y-%m-%d")

            # 招商员分配
            if is_douyin_kezi:
                agent = get_val(row, agent_col) or "郑建军"
            else:
                agent = (
                    get_val(row, find_col(["所属招商"]))
                    or get_val(row, find_col(["跟进员工"]))
                    or get_val(row, agent_col)
                    or "郑建军"
                )

            parsed.append(
                {
                    "phone": phone,
                    "platform": "小红书"
                    if is_xhs_channel
                    else (
                        "抖音"
                        if is_douyin_channel
                        else (get_val(row, platform_col) or "抖音")
                    ),
                    "agent": agent,
                    "entry_date": entry_date,
                    "name": get_val(row, name_col),
                    "city": get_val(row, city_col),
                    "region": get_val(row, region_col),
                    "validity": get_val(row, validity_col),
                    "can_wechat": get_val(row, wechat_col),
                    "remark": get_val(row, remark_col),
                    "follow_time": parse_date(row, follow_time_col),
                    "follow_note": get_val(row, follow_note_col),
                    "call_time": parse_date(row, call_time_col),
                    "visit_time": parse_date(row, visit_time_col),
                    "sign_time": parse_date(row, sign_time_col),
                    # 小红书客资专用字段
                    "xhs_account": get_val(row, xhs_account_col) if is_xhs_channel else "",
                    "lead_type": get_val(row, lead_type_col) if is_xhs_channel else "",
                }
            )

        if not parsed:
            return jsonify({"success": False, "message": "未能解析出任何有效数据"})

        print(f"[导入] 解析成功: {len(parsed)} 条, 空值跳过: {len(bad_rows)} 条")

        # ── 5) 数据库写入 ──
        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()

        # 5.1) 清理已有重复
        c.execute(
            "SELECT phone, COUNT(*) as cnt FROM new_leads GROUP BY phone HAVING cnt > 1"
        )
        for phone, cnt in c.fetchall():
            c.execute(
                "DELETE FROM new_leads WHERE phone = ? AND id NOT IN (SELECT MIN(id) FROM new_leads WHERE phone = ?)",
                (phone, phone),
            )
        conn.commit()

        # 5.2) 加载已有线索
        c.execute("SELECT phone, platform, entry_date FROM new_leads")
        existing = {row[0]: {"platform": row[1], "entry_date": row[2]} for row in c.fetchall()}

        added, updated, dup_skip, skipped = 0, 0, 0, 0
        seen_in_file = set()

        for item in parsed:
            phone = item["phone"]
            if phone in seen_in_file:
                dup_skip += 1
                continue
            seen_in_file.add(phone)

            platform = item["platform"]
            agent = item["agent"]
            entry_date = item["entry_date"]

            if phone in existing:
                old = existing[phone]
                # 抖音/小红书渠道导入时，已存在的线索直接跳过（由管理表维护）
                if is_douyin_channel or is_xhs_channel:
                    skipped += 1
                    continue

                # 招商线索管理表导入时，抖音/小红书入库日期不变（兼容"抖音广告"等变体）
                if is_zhaoshang and (
                    "抖音" in old["platform"] or "小红书" in old["platform"]
                ):
                    entry_date = old["entry_date"]

                # UPDATE 全部字段（region 加入更新）
                c.execute(
                    """UPDATE new_leads SET
                    platform = ?, agent = ?, entry_date = ?, name = ?,
                    city = ?, region = ?, validity = ?, can_wechat = ?, remark = ?,
                    二次联系时间 = ?, 二次联系备注 = ?, 最近一次电联时间 = ?, 到访时间 = ?, 签约时间 = ?,
                    xhs_account = ?, lead_type = ?
                    WHERE phone = ?""",
                    (
                        platform,
                        agent,
                        entry_date,
                        item["name"],
                        item["city"],
                        item["region"],
                        item["validity"],
                        item["can_wechat"],
                        item["remark"],
                        item["follow_time"],
                        item["follow_note"],
                        item["call_time"],
                        item["visit_time"],
                        item["sign_time"],
                        item["xhs_account"],
                        item["lead_type"],
                        phone,
                    ),
                )
                updated += 1
            else:
                # 招商线索管理表导入时，抖音/小红书的新线索不创建（由渠道表格单独导入）
                if is_zhaoshang and ("抖音" in platform or "小红书" in platform):
                    skipped += 1
                    continue

                # INSERT 新线索（region 加入插入）
                c.execute(
                    """INSERT INTO new_leads
                    (phone, platform, agent, entry_date, name, city, region, validity,
                     can_wechat, remark, created_at,
                     二次联系时间, 二次联系备注, 最近一次电联时间, 到访时间, 签约时间,
                     xhs_account, lead_type)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        phone,
                        platform,
                        agent,
                        entry_date,
                        item["name"],
                        item["city"],
                        item["region"],
                        item["validity"],
                        item["can_wechat"],
                        item["remark"],
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        item["follow_time"],
                        item["follow_note"],
                        item["call_time"],
                        item["visit_time"],
                        item["sign_time"],
                        item["xhs_account"],
                        item["lead_type"],
                    ),
                )
                added += 1
                existing[phone] = {"platform": platform, "entry_date": entry_date}

        conn.commit()
        conn.close()

        msg = f"导入完成！新增 {added} 条，更新 {updated} 条"
        if skipped:
            if is_douyin_channel or is_xhs_channel:
                msg += f"，已存在线索跳过 {skipped} 条"
            else:
                msg += f"，抖音/小红书新线索跳过 {skipped} 条（请通过渠道表格导入）"
        if dup_skip:
            msg += f"，Excel内重复跳过 {dup_skip} 条"
        if bad_rows:
            msg += f"，空值跳过 {len(bad_rows)} 条"

        return jsonify(
            {
                "success": True,
                "message": msg,
                "added": added,
                "updated": updated,
                "dup_skip": dup_skip,
                "bad": len(bad_rows),
                "bad_rows": bad_rows[:5],
            }
        )

    except Exception as e:
        import traceback

        print(f"[导入错误] {str(e)}\n{traceback.format_exc()}")
        return jsonify({"success": False, "message": f"导入失败: {str(e)}"})


@leads_bp.route("/api/leads/import-douyin", methods=["POST"])
def import_douyin_kezi():
    user = session.get("user")
    if not user or user["role"] != "admin":
        return jsonify({"success": False, "message": "只有管理员可以导入"}), 401

    if "file" not in request.files:
        return jsonify({"success": False, "message": "请选择文件"})

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"success": False, "message": "请选择文件"})

    try:
        import pandas as pd
        from io import BytesIO
        from openpyxl import load_workbook

        file_bytes = file.read()
        bio = BytesIO(file_bytes)
        filename_lower = file.filename.lower()
        is_xls = filename_lower.endswith(".xls")

        # 获取 sheet 名称列表
        if is_xls:
            xls = pd.ExcelFile(bio, engine="xlrd")
            sheet_names = xls.sheet_names
        else:
            bio.seek(0)
            wb = load_workbook(bio, data_only=True)
            sheet_names = wb.sheetnames

        def load_sheet(bio, filename, sheet_name=None):
            """统一读取Excel返回DataFrame。xlsx走openpyxl，xls走pandas+xlrd"""
            if filename.lower().endswith(".xls"):
                bio.seek(0)
                if sheet_name:
                    df = pd.read_excel(bio, sheet_name=sheet_name, engine="xlrd")
                else:
                    df = pd.read_excel(bio, engine="xlrd")
            else:
                bio.seek(0)
                wb = load_workbook(bio, data_only=True)
                ws = wb[sheet_name] if sheet_name else wb.active
                data = []
                headers = None
                for row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(c).strip() if c is not None else "" for c in row]
                    else:
                        data.append(row)
                df = pd.DataFrame(data, columns=headers)
            if df is not None and len(df) > 0:
                df = df.dropna(how="all")
            return df

        # 直接读取第一个有数据的sheet
        df = None
        debug_info = []
        try:
            df = load_sheet(bio, file.filename)
            debug_info.append(f"直接读取成功，行数={len(df)}, 列数={len(df.columns)}")
        except Exception as e:
            debug_info.append(f"直接读取失败: {e}")

        # 如果直接读取为空，逐个sheet尝试
        if df is None or len(df) == 0:
            for sheet_name in sheet_names:
                try:
                    df = load_sheet(bio, file.filename, sheet_name)
                    debug_info.append(f"Sheet {sheet_name} 读取成功，行数={len(df)}")
                    if len(df) > 0:
                        break
                except Exception as e2:
                    debug_info.append(f"Sheet {sheet_name} 读取失败: {e2}")
                    continue

        if df is None or len(df) == 0:
            return jsonify(
                {
                    "success": False,
                    "message": f'Excel 文件为空。调试: {" | ".join(debug_info)}',
                }
            )

        cols = [str(c).strip() for c in df.columns]

        # 查找对应列
        def find_col(keywords):
            for kw in keywords:
                kw_lower = kw.lower()
                for col in cols:
                    if kw_lower in col.lower():
                        return col
            return None

        date_col = find_col(["线索创建时间", "创建时间", "日期", "时间"])
        name_col = find_col(["姓名", "名字", "客户姓名"])
        phone_col = find_col(["手机号", "手机号码", "电话", "联系电话", "手机"])
        agent_col = find_col(["跟进员工", "所属招商", "负责人", "招商员", "员工"])
        city_col = find_col(["所在城市", "城市", "省份", "地区"])

        if not phone_col:
            return jsonify(
                {"success": False, "message": f"无法识别手机号列。当前列名: {cols}"}
            )

        # 加载已有手机号（用于去重）
        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()
        c.execute("SELECT phone FROM new_leads")
        existing_phones = {row[0] for row in c.fetchall()}
        conn.close()

        def get_val(row, col):
            if not col:
                return ""
            v = row.get(col, "")
            if pd.isna(v):
                return ""
            s = str(v).strip()
            return s if s.lower() != "nan" else ""

        def parse_date(row, col):
            if not col:
                return ""
            v = row.get(col)
            if pd.isna(v):
                return ""
            try:
                return pd.to_datetime(v).strftime("%Y-%m-%d")
            except (ValueError, TypeError):
                return ""

        added, skipped, bad = 0, 0, 0
        seen = set()

        conn = sqlite3.connect(str(DB_FILE))
        c = conn.cursor()

        for idx, row in df.iterrows():
            raw_phone = get_val(row, phone_col)
            if not raw_phone:
                bad += 1
                continue

            # 提取数字
            digits = "".join(filter(str.isdigit, raw_phone))
            if len(digits) >= 7:
                phone = digits
            else:
                phone = raw_phone

            # 去重：数据库已有 或 Excel内重复
            if phone in existing_phones or phone in seen:
                skipped += 1
                continue
            seen.add(phone)

            entry_date = parse_date(row, date_col) or datetime.now().strftime("%Y-%m-%d")
            name = get_val(row, name_col)
            agent = get_val(row, agent_col) or "郑建军"
            city = get_val(row, city_col)

            c.execute(
                """INSERT INTO new_leads
                (phone, platform, agent, entry_date, name, city, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (
                    phone,
                    "抖音",
                    agent,
                    entry_date,
                    name,
                    city,
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )
            added += 1

        conn.commit()
        conn.close()

        return jsonify(
            {
                "success": True,
                "message": f"导入完成！新增 {added} 条，重复跳过 {skipped} 条，无法识别 {bad} 条",
                "added": added,
                "skipped": skipped,
                "bad": bad,
            }
        )

    except Exception as e:
        import traceback

        print(f"[抖音导入错误] {str(e)}\n{traceback.format_exc()}")
        return jsonify({"success": False, "message": f"导入失败: {str(e)}"})
